import os
import io
import json
import hashlib
from flask import Flask, render_template, jsonify, request, send_file
from models import get_db, init_db
from game_logic import (
    get_current_state, validate_selection, record_selection, reset_game,
    is_game_over, get_snake_order,
    init_pin_for_participant, verify_pin_for_participant,
    get_queue_for_participant, save_queue_for_participant,
    toggle_auto_draft_for_participant, get_auto_draft_state_for_participant,
    _execute_auto_draft
)
from seed_data import seed_database, sync_team_ownership, TEAM_INFO_PATH
from models import migrate_briefing_tables
from briefing_data import (
    load_briefing,
    load_briefing_enriched,
    load_history_index,
    beijing_date_label,
    kickoff_beijing_label,
    get_report_for_date,
    history_dates_payload,
    get_match_detail,
    get_owner_map,
    get_selections_for_display,
    save_json,
    LATEST_PATH,
    HISTORY_PATH,
    FIXTURES_PATH,
    SQUAD_META_PATH,
    TEAM_FORM_PATH,
)
from briefing.rebuild_scorers import rebuild_scorers_from_api
from briefing.scorer_match import add_manual_rule
from briefing.shooter_standings import (
    STANDINGS_PATH as SHOOTER_STANDINGS_PATH,
    load_shooter_standings,
)
from briefing.standings import STANDINGS_PATH, compute_team_standings, load_team_standings

PARTICIPANT_COLORS = {
    '耗子': '#f5c518',
    '庆爷': '#4ade80',
    '李总': '#38bdf8',
    '老闫': '#f87171',
    '老王': '#c084fc',
}
import openpyxl

app = Flask(__name__)


@app.template_filter('beijing_date')
def _filter_beijing_date(iso_date):
    return beijing_date_label(iso_date)


@app.template_filter('kickoff_beijing')
def _filter_kickoff_beijing(kickoff):
    return kickoff_beijing_label(kickoff)


def get_teams_with_players():
    db = get_db()
    teams = db.execute('''
        SELECT t.id, t.name, t.group_name, t.flag_emoji
        FROM teams t ORDER BY t.group_name, t.id
    ''').fetchall()

    # Get selected player IDs
    selected_ids = set(
        r['player_id'] for r in db.execute('SELECT player_id FROM selections').fetchall()
    )

    result = []
    for t in teams:
        players = db.execute(
            'SELECT id, name, name_cn, jersey_number, position FROM players WHERE team_id = ? ORDER BY jersey_number',
            (t['id'],)
        ).fetchall()
        result.append({
            'id': t['id'],
            'name': t['name'],
            'group_name': t['group_name'],
            'flag_emoji': t['flag_emoji'],
            'players': [{
                'id': p['id'],
                'name': p['name'],
                'name_cn': p['name_cn'] or p['name'],
                'jersey_number': p['jersey_number'],
                'position': p['position'],
                'selected': p['id'] in selected_ids,
            } for p in players],
        })
    db.close()
    return result


# ========== Routes ==========

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/stats')
def stats():
    import os
    stats_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'wcq_player_stats_table.html')
    return send_file(stats_path)


@app.route('/teams')
def teams():
    import os
    teams_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'team_ranking_top500.html')
    return send_file(teams_path)


@app.route('/briefing')
def briefing_page():
    latest = load_briefing_enriched()
    hist = history_dates_payload()
    report_date = request.args.get('report_date') or hist.get('default') or ''
    return render_template(
        'briefing.html',
        latest=latest,
        hist=hist,
        report_date=report_date,
    )


@app.route('/match/<int:fixture_id>')
def match_page(fixture_id):
    detail = get_match_detail(fixture_id)
    if not detail:
        return '比赛未找到', 404
    return render_template('match_intro.html', m=detail)


@app.route('/standings/teams')
def standings_teams_page():
    return render_template(
        'standings_teams.html',
        standings=load_team_standings(),
        colors=PARTICIPANT_COLORS,
    )


@app.route('/api/standings/teams')
def api_standings_teams():
    return jsonify(load_team_standings())


@app.route('/standings/shooters')
def standings_shooters_page():
    return render_template(
        'standings_shooters.html',
        standings=load_shooter_standings(),
        colors=PARTICIPANT_COLORS,
        selections=get_selections_for_display(),
    )


@app.route('/api/standings/shooters')
def api_standings_shooters():
    return jsonify(load_shooter_standings())


@app.route('/api/standings/shooters/repair', methods=['POST'])
def api_standings_shooters_repair():
    token = request.args.get('token') or request.headers.get('X-Import-Token', '')
    expected = os.environ.get('IMPORT_BRIEFING_TOKEN', '')
    if not expected or token != expected:
        return jsonify({'success': False, 'error': 'unauthorized'}), 403
    data = request.get_json(force=True, silent=True) or {}
    player_id = data.get('player_id')
    if not player_id:
        return jsonify({'success': False, 'error': 'player_id required'}), 400
    add_manual_rule(
        player_id=int(player_id),
        api_scorer_en=data.get('api_scorer_en'),
        team_api=data.get('team_api'),
        api_scorer_id=data.get('api_scorer_id'),
        note=data.get('note', '后台修复'),
    )
    result = rebuild_scorers_from_api()
    return jsonify({'success': True, 'standings': result.get('standings')})


@app.route('/api/briefing')
def api_briefing():
    data = load_briefing_enriched()
    if not data:
        return jsonify({'error': 'no briefing data'}), 404
    return jsonify(data)


@app.route('/api/briefing/history/dates')
def api_briefing_history_dates():
    return jsonify(history_dates_payload())


@app.route('/api/briefing/history/<date>')
def api_briefing_history_date(date):
    report = get_report_for_date(date)
    if not report:
        return jsonify({'error': 'not found'}), 404
    return jsonify(report)


@app.route('/api/match/<int:fixture_id>')
def api_match(fixture_id):
    detail = get_match_detail(fixture_id)
    if not detail:
        return jsonify({'error': 'not found'}), 404
    return jsonify(detail)


@app.route('/api/team-ownership')
def api_team_ownership():
    owners = get_owner_map()
    return jsonify(owners)


@app.route('/api/import-briefing', methods=['POST'])
def api_import_briefing():
    from briefing.validate import validate_briefing_payload

    token = request.args.get('token') or request.headers.get('X-Import-Token', '')
    expected = os.environ.get('IMPORT_BRIEFING_TOKEN', '')
    if not expected or token != expected:
        return jsonify({'success': False, 'error': 'unauthorized'}), 403
    data = request.get_json(force=True, silent=True) or {}
    ok, errors = validate_briefing_payload(data)
    if not ok:
        return jsonify({'success': False, 'error': 'validation failed', 'details': errors}), 400
    if 'latest' in data:
        save_json(LATEST_PATH, data['latest'])
    if 'history_index' in data:
        save_json(HISTORY_PATH, data['history_index'])
    if 'standings_teams' in data:
        save_json(STANDINGS_PATH, data['standings_teams'])
    if 'standings_shooters' in data:
        save_json(SHOOTER_STANDINGS_PATH, data['standings_shooters'])
    if 'team_squad_meta' in data:
        save_json(SQUAD_META_PATH, data['team_squad_meta'])
    if 'team_form' in data:
        save_json(TEAM_FORM_PATH, data['team_form'])
    if 'fixtures_2026' in data:
        save_json(FIXTURES_PATH, data['fixtures_2026'])
    return jsonify({'success': True})


@app.route('/api/state')
def api_state():
    state = get_current_state()
    return jsonify(state)


@app.route('/api/teams')
def api_teams():
    return jsonify(get_teams_with_players())


@app.route('/api/players')
def api_players():
    return jsonify(get_teams_with_players())


@app.route('/api/select', methods=['POST'])
def api_select():
    data = request.get_json(force=True)
    participant_name = data.get('participant_name')
    player_id = data.get('player_id')
    pin = data.get('pin', '')

    if not participant_name or not player_id:
        return jsonify({'success': False, 'error': '\u7f3a\u5c11\u53c2\u6570'}), 400

    # Verify PIN before allowing selection
    if not verify_pin_for_participant(participant_name, pin):
        return jsonify({'success': False, 'error': 'PIN\u9a8c\u8bc1\u5931\u8d25'}), 403

    validation = validate_selection(participant_name, player_id)
    if not validation['valid']:
        return jsonify({'success': False, 'error': validation['error']}), 400

    auto_picks = record_selection(participant_name, player_id)
    return jsonify({'success': True, 'auto_picks': auto_picks, 'state': get_current_state()})


@app.route('/api/selections')
def api_selections():
    state = get_current_state()
    return jsonify(state['selections'])


@app.route('/api/export')
def api_export():
    db = get_db()
    # Build grid: rows = rounds, cols = participants
    participants = db.execute(
        'SELECT id, name, draft_order FROM participants ORDER BY draft_order'
    ).fetchall()

    selections = db.execute('''
        SELECT s.round_number, s.pick_number,
               p.name AS participant_name, pl.name AS player_name,
               pl.name_cn AS player_name_cn,
               pl.jersey_number, t.name AS team_name
        FROM selections s
        JOIN participants p ON s.participant_id = p.id
        JOIN players pl ON s.player_id = pl.id
        JOIN teams t ON pl.team_id = t.id
        ORDER BY s.round_number, s.pick_number
    ''').fetchall()

    # Build round × participant grid
    grid = {}
    for s in selections:
        r = s['round_number']
        order = None
        for p in participants:
            if p['name'] == s['participant_name']:
                order = p['draft_order']
                break
        if order:
            col = order + 1  # col 1=round, col 2-6=participants by draft_order
            if r not in grid:
                grid[r] = {}
            name = s['player_name_cn'] or s['player_name']
            grid[r][col] = f"{name} #{s['jersey_number']} ({s['team_name']})"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '选秀结果'

    # Header
    ws.cell(1, 1, '轮次')
    for i, p in enumerate(participants):
        ws.cell(1, i + 2, p['name'])

    # Data
    max_round = max(grid.keys()) if grid else 0
    for r in range(1, max_round + 1):
        ws.cell(r + 1, 1, r)
        for cp in range(2, 7):
            if r in grid and cp in grid[r]:
                ws.cell(r + 1, cp, grid[r][cp])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    db.close()
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='选秀结果.xlsx')


@app.route('/api/reset', methods=['POST'])
def api_reset():
    seed_database()
    return jsonify({'success': True})


@app.route('/api/preselect/init-pin', methods=['POST'])
def api_init_pin():
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({'success': False, 'error': '无法解析JSON: ' + str(request.data)[:200]}), 400
    name = data.get('participant_name')
    pin = data.get('pin')
    if not name or not pin:
        return jsonify({'success': False, 'error': '缺少参数: name=' + repr(name) + ' pin=' + repr(pin)}), 400
    if init_pin_for_participant(name, pin):
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'PIN已存在'}), 400


@app.route('/api/preselect/verify-pin', methods=['POST'])
def api_verify_pin():
    data = request.get_json()
    name = data.get('participant_name')
    pin = data.get('pin')
    if not name or not pin:
        return jsonify({'success': False, 'error': '缺少参数'}), 400
    return jsonify({'success': verify_pin_for_participant(name, pin)})


@app.route('/api/preselect/<name>')
def api_get_preselect(name):
    pin = request.args.get('pin', '')
    db = get_db()
    row = db.execute(
        'SELECT pin_hash, queue_data FROM preselect_queues WHERE participant_name = ?',
        (name,)
    ).fetchone()

    if not row:
        db.close()
        return jsonify({'queue': [], 'has_pin': False})

    if not pin:
        db.close()
        # Only report has_pin if hash is actually set (not just row exists from auto-draft)
        has_pin = bool(row['pin_hash'])
        return jsonify({'queue': [], 'has_pin': has_pin})

    # Verify PIN
    if hashlib.sha256(pin.encode()).hexdigest() != row['pin_hash']:
        db.close()
        return jsonify({'queue': [], 'has_pin': True, 'pin_ok': False}), 403

    # Return queue filtered by already-selected
    selected_ids = set(
        r['player_id'] for r in db.execute('SELECT player_id FROM selections').fetchall()
    )
    db.close()
    queue = json.loads(row['queue_data'])
    filtered = [pid for pid in queue if pid not in selected_ids]
    return jsonify({'queue': filtered, 'has_pin': True, 'pin_ok': True})


@app.route('/api/preselect/<name>', methods=['POST'])
def api_save_preselect(name):
    data = request.get_json(force=True)
    pin = data.get('pin', '')
    queue = data.get('queue', [])
    if save_queue_for_participant(name, pin, queue):
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'PIN错误'}), 403


@app.route('/api/auto-draft/toggle', methods=['POST'])
def api_toggle_auto_draft():
    data = request.get_json(force=True)
    name = data.get('participant_name')
    enabled = data.get('enabled', False)
    if not name:
        return jsonify({'success': False, 'error': '缺少参数'}), 400
    toggle_auto_draft_for_participant(name, enabled)
    # On enabling, immediately check if it's this participant's turn
    auto_picks = []
    if enabled:
        db = get_db()
        auto_picks = _execute_auto_draft(db)
        db.commit()
        db.close()
    return jsonify({'success': True, 'auto_picks': auto_picks})


@app.route('/api/auto-draft/state/<name>')
def api_auto_draft_state(name):
    return jsonify({'enabled': get_auto_draft_state_for_participant(name)})


# ========== Init ==========

def ensure_db_initialized():
    """Initialize and seed DB on first run, re-seed if name_cn is missing"""
    migrate_briefing_tables()
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'draft.db')
    if not os.path.exists(db_path):
        seed_database()
        _sync_ownership_if_needed()
        return
    db = get_db()
    empty_count = db.execute('SELECT COUNT(*) FROM players WHERE name_cn = ""').fetchone()[0]
    own_count = db.execute('SELECT COUNT(*) FROM team_ownership').fetchone()[0]
    db.close()
    if empty_count > 0:
        seed_database()
    if own_count != 40:
        _sync_ownership_if_needed()


def _sync_ownership_if_needed():
    if os.path.isfile(TEAM_INFO_PATH):
        try:
            sync_team_ownership()
        except (ValueError, FileNotFoundError) as e:
            print(f'team_ownership sync skipped: {e}')


ensure_db_initialized()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
